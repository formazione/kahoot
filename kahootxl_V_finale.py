import openpyxl as opx
import os
from random import shuffle


wb = opx.load_workbook("esempio.xlsx")
ws = wb.active

'''
etichetta
pacchetto igiene
tu salute e sicurezza sul lavoro

Regolamento UE 1169/2011
d.lgs. 193/2007
d.lgs 81/2008
art. 1321 c.c.


'''


# put a space for empty answers
data1 = """
Il d.lgs. 193/2007 riguarda le norme su...
igiene
lavoro
antincendio
privacy
30

Il controllo interno è quello fatto tramite
HACCP
NAS
ASL
60

Tracciabilità =
dalla fattoria alla tavola
dalla tavola alla fattoria
10

L'etichettatura è regolata da
Regolamento UE 1169/2011
d.lgs. 193/2007
d.lgs 81/2008
art. 1321 c.c.
20

La definizione del contratto:
art. 1321 c.c.
art. 2555 c.c.
art. 2082 c.c.
art. 2424 c.c.
20

DVR (d.lgs. 81/2008) significa:
documento di valutazione dei rischi
documento di verifica ricavi
documento validazione regolamento
30

L'HACCP fu creato da:
NASA
NATO
ONU
OMT
20

Il TU sulla salute e sicurezza sul lavoro è stato emanato con:
dlgs 81/2008
dm 9 aprile 1994
dlgs 6 novembre 2007
art. 1321 c.c.
20

Il datore di lavoro nomina per il dlgs 81/2008:
il medico competente
il rappresentante della sicurezza dei lavoratori
15

Nome dell'impresa
Ditta
Marchio
Insegna
L'etichetta
20

L'insegna è un segno distintivo
facoltativo
obbligatorio

Tutte le fasi di lavorazione si eseguono sul territorio:
DOP
IGP
STG

L'Arancia rossa d Sicilia è:
IGP
DOP
STG

La mozzarella è
STG
DOP
IGP

Si può produrre anche in altri territorio il prodotto:
STG
DOP
DOC
IGP

Il Provolone Valpadana è:
DOP
DOC
IGP
STG"""

data2 = data1[1:].split("\n\n")
print(f"NUMERO DOMANDE: {len(data2)}")
data3 = []
for d in data2: # ['\nQuestion\n1\n2\n3\n4\n30', 'Question\n1\n2\n3\n4\n30\n']
	data3.append(d.split("\n"))
print(data3)

data4 = []
for d in data3:
	# [['Question', '1', '2', '3', '4', '30'], ['Question', '10', '20', '30', '40', '60']]
	print(type(d[-2]))
	if not d[-1].isdigit():
		d.append(20)
	if len(d) == 6:
		data4.append([d[0], ",".join([d[1],d[2],d[3],d[4]]), d[5]])
	if len(d) == 5:
		data4.append([d[0], ",".join([d[1],d[2],d[3]]), d[4]])
	if len(d) == 4:
		data4.append([d[0], ",".join([d[1],d[2]]), d[3]])

# data4 = [ ["question", "risposte,risposte,risposte", tempo]
print(data4)

count = 9
for d in data4:
	print(count)
	correct = d[1].split(",")[0] # la prima è quella esatta
	rnd = d[1].split(",") # separa le risposte in una lista
	numero_risposte = len(rnd)
	shuffle(rnd) # mischia le risposte
	print(rnd)
	correct = rnd.index(correct) + 1 # trova l'indice della corretta
	match numero_risposte: # a seconda del numero di risposte
		case 4:
			r1, r2, r3, r4 = rnd
			# inserisce nelle colonne le risposte
			ws[f"C{count}"] = r1
			ws[f"D{count}"] = r2
			ws[f"E{count}"] = r3
			ws[f"F{count}"] = r4
		case 3:
			r1, r2, r3 = rnd
			ws[f"C{count}"] = r1
			ws[f"D{count}"] = r2
			ws[f"E{count}"] = r3
		case 2:
			r1, r2 = rnd
			ws[f"C{count}"] = r1
			ws[f"D{count}"] = r2

	ws[f"B{count}"] = d[0] # domanda nella colonna B
	ws[f"G{count}"] = d[2] # TEMPO per rispondere
	ws[f"H{count}"] = correct # numero risposta esatta
	count += 1

wb.save("esempio.xlsx")
os.startfile("esempio.xlsx")