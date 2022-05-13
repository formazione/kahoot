import openpyxl as opx
import os
from random import shuffle


wb = opx.load_workbook("esempio.xlsx")
ws = wb.active

'''
etichetta
pacchetto igiene
tu salute e sicurezza sul lavoro

Regolamento UE 1169/2011
d.lgs. 193/2007
d.lgs 81/2008
art. 1321 c.c.


'''


# put a space for empty answers
data1 = """
Il d.lgs. 193/2007 riguarda le norme su...
igiene
lavoro
antincendio
privacy
30

Il controllo interno è quello fatto tramite
HACCP
NAS
ASL

60

Tracciabilità =
dalla fattoria alla tavola
dalla tavola alla fattoria
 

10

L'etichettatura è regolata da
Regolamento UE 1169/2011
d.lgs. 193/2007
d.lgs 81/2008
art. 1321 c.c.
20

La definizione del contratto:
art. 1321 c.c.
art. 2555 c.c.
art. 2082 c.c.
art. 2424 c.c.
20

DVR (d.lgs. 81/2008) significa:
documento di valutazione dei rischi
documeto di verifica ricavi
documento validazione regolamento
 
30

L'HACCP fu creato da:
NASA
NATO
ONU
OMT
20
"""

data2 = data1[1:].split("\n\n")

data3 = []
for d in data2: # ['\nQuestion\n1\n2\n3\n4\n30', 'Question\n1\n2\n3\n4\n30\n']
	data3.append(d.split("\n"))
print(data3)

data4 = []
for d in data3:
	# [['Question', '1', '2', '3', '4', '30'], ['Question', '10', '20', '30', '40', '60']]
	if len(d) == 6:
		data4.append([d[0], ",".join([d[1],d[2],d[3],d[4]]), d[5]])
	if len(d) == 5:
		data4.append([d[0], ",".join([d[1],d[2],d[3]]), d[4]])
	if len(d) == 4:
		data4.append([d[0], ",".join([d[1],d[2]]), d[3]])

print(data4)

count = 9
for d in data4:
	print(count)
	correct = d[1].split(",")[0] 
	rnd = d[1].split(",")
	lenr = len(rnd)
	shuffle(rnd)
	print(rnd)
	correct = rnd.index(correct) + 1
	match lenr:
		case 4:
			r1, r2, r3, r4 = rnd

			ws[f"C{count}"] = r1
			ws[f"D{count}"] = r2
			ws[f"E{count}"] = r3
			ws[f"F{count}"] = r4
		case 3:
			r1, r2, r3 = rnd
			ws[f"C{count}"] = r1
			ws[f"D{count}"] = r2
			ws[f"E{count}"] = r3
		case 2:
			r1, r2 = rnd
			ws[f"C{count}"] = r1
			ws[f"D{count}"] = r2

	ws[f"B{count}"] = d[0]
	ws[f"G{count}"] = d[2]
	ws[f"H{count}"] = correct
	count += 1

wb.save("esempio.xlsx")
os.startfile("esempio.xlsx")